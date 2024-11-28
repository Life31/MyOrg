from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from .forms import PassForm, Car_passForm, PassForm_Admin
from .models import Pass, Car_pass
from django.conf import settings
#from win32com import client
import openpyxl
import shutil
import os
#import pythoncom
import transliterate
import datetime as dt
from django.db.models import Q

rights = {'rights': settings.RIGHTS,
          'storage_rights': settings.STORAGE_RIGHTS,
          'sadec_rights': settings.SADEC_RIGHTS,
          'pro_rights': settings.PRO_RIGHTS,
          'passes_rights': settings.PASSES_RIGHTS, }

urls = {
    'passes_urls': {
        'all_passes', 'pass_new', 'pass_delete',
        'pass_name_filtr', 'pass_edit', 'pass_print',
    },
}


# Главная страница
def all_passes(request):
    passes = Pass.objects.all()
    paginator = Paginator(passes, 20)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)
    return render(request,
                  'all_passes.html',
                  {'page': page, **rights, **urls}
                  )


@login_required
def pass_new(request):
    pas = Pass()
    try:
        pas.num =  int(Pass.objects.all().latest('id').num) + 1
    except:
        pas.num = 1
    pas.author = request.user.username
    form = PassForm(request.POST or None,
                        files=request.FILES or None,
                        instance=pas)

    if form.is_valid():
        form.save()
        #pass_print(pas)   
        return redirect('all_passes')
    return render(request, 'pass_new.html', {'form': form, **rights, **urls})


@login_required
def pass_copy(request, pass_id):
    try:
        pas = get_object_or_404(Pass, id=pass_id)
        form = PassForm(request.POST or None,
                            files=request.FILES or None,
                            instance=pas)

        if form.is_valid():
            new_pas = form.save(commit=False)
            new_pas.id = int(Pass.objects.all().latest('id').id) + 1
            try:
                new_pas.num =  int(Pass.objects.all().latest('id').num) + 1
            except:
                new_pas.num = 1
            new_pas.author = request.user.username
            new_pas.save()
            #pass_print(new_pas)
            return redirect('all_passes')
    except:
        return redirect('all_passes')
    return render(request, 'pass_new.html', {'form': form, **rights, **urls})

@login_required
def pass_delete(request, pass_id):
    if request.user.username not in settings.RIGHTS:
        return redirect('all_passes')
    pas = Pass.objects.filter(id=pass_id)[0]
    print(pas.file)

    try:
        os.remove(settings.MEDIA_ROOT + pas.file.url)
        print(f'Файл {settings.MEDIA_ROOT + pas.file.url} успешно удален')
    except:
        print(f'Не удалось удалить файл {settings.MEDIA_ROOT + "/" + pas.file.url}')
    pas.delete()

    return redirect('all_passes')


def pass_name_filtr(request, sec_name):
    passes = Pass.objects.filter(sec_name=sec_name)
    paginator = Paginator(passes, 10)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)
    content = {'page': page,}
    return render(request, 'all_passes.html', {**content, **rights, **urls })

def pass_search(request):
    q = request.GET.get("q")
    passes = Pass.objects.filter(
        Q(sec_name__icontains=q)
        | Q(name__icontains=q)
        | Q(patro__icontains=q)
        | Q(where__icontains=q) 
    )     
    #passes = Pass.objects.filter(sec_name__icontains=q)
    #paginator = Paginator(passes, 10)
    #page_number = request.GET.get('page')
    #page = paginator.get_page(page_number)
    content = {'page': passes} 
    return render(request, 'all_passes.html', {**content, **rights, **urls })


@login_required
def pass_edit(request, pass_id):
    pas = Pass.objects.filter(id=pass_id)[0]
    pas.day = str(pas.day)[:-15]
    #if request.user == pas.author or request.user.username in settings.RIGHTS:
    form = PassForm(request.POST or None,
                     files=request.FILES or None,
                    instance=pas)
    if form.is_valid():
        form.save()
        #pass_print(pas)
        return redirect('all_passes')
    return render(request, 'pass_new.html', {'form': form, 'edit': "edit", **rights, **urls})


#@login_required
def pass_print(pas):

    monthes = { 1: 'января',
                2: 'февраля',
                3: 'марта',
                4: 'апреля',
                5: 'мая',
                6: 'июня',
                7: 'июля',
                8: 'августа',
                9: 'сентября',
                10: 'октября',
                11: 'ноября',
                12: 'декабря', }
    print(settings.BASE_DIR)
    sample = settings.BASE_DIR + "\posts\static\passes\pass_doc.xlsx"
    #direct = "storage/"
    #file_name = sample[:-14] + str(post.title) + ".xlsx"        
    #shutil.copy2(sample, file_name)        
    wb = openpyxl.load_workbook(sample)
    ws = wb['Пропуска']
    ws['B7'] = pas.sec_name
    ws['B8'] = pas.name
    ws['B9'] = pas.patro
    ws['D9'] = pas.sec_name
    ws['D10'] = pas.name + ' ' + pas.patro
    ws['D13'] = pas.where
    ws['D15'] = pas.doc_type
    ws['D16'] = pas.pasport
    ws['F18'] = pas.spec
    ws['H1'] = pas.sec_name + ' ' + pas.name + ' ' + pas.patro
    ws['D7'] = pas.day.day
    ws['E7'] = monthes[pas.day.month]
    ws['F7'] = str(pas.day.year) + ' ' + 'года'

    wb.save(sample) # Сохраняем измененный файл
    #os.startfile(sample)
    # Open Microsoft Excel
    #pythoncom.CoInitializeEx(0)
    #excel = client.Dispatch("Excel.Application")
    #excel.Visible = False
    # Read Excel File
    #sheets = excel.Workbooks.Open(sample)
    #work_sheets = sheets.Worksheets[0]

    # Convert into PDF File


    try:
        pdf_file_name = (
            settings.BASE_DIR
            + "\posts\static\passes\\"
            + transliterate.translit(
                str(pas.sec_name) + "_"
                + str(pas.name) + "_"
                + str(pas.patro) + "("
                + str(pas.num) + ").pdf", reversed=True
            )
        )
    except:
        pdf_file_name = (settings.BASE_DIR +
                         "\posts\static\passes\\" +
                         str(pas.sec_name) + "_" +
                         str(pas.name) + "_" +
                         str(pas.patro) + "(" +
                         str(pas.num) + ").pdf")
    #work_sheets.ExportAsFixedFormat(0, pdf_file_name)
    pas.file = pdf_file_name[pdf_file_name.rfind('passes\\'):]

    pas.save()
    #work_sheets.ExportAsFixedFormat(0, settings.BASE_DIR + "\posts\static\passes\pass_doc.pdf")
    #sheets.Close(True)
    #post.cz = direct + str(post.title) + ".xlsx"
