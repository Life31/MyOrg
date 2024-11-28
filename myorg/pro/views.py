from django.shortcuts import render
from .models import Block, Event  #, Block_state
from .forms import BlockForm, EventForm
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.db.models import Q
#from datetime import timedelta, date
import datetime as dt
from django.conf import settings
import openpyxl
from openpyxl.styles import  Border, Side, PatternFill, Alignment #, Font
#import pythoncom
#import transliterate
#from win32com import client


rights = {
    'rights': settings.RIGHTS,
    'storage_rights': settings.STORAGE_RIGHTS,
    'sadec_rights': settings.SADEC_RIGHTS,
    'pro_rights': settings.PRO_RIGHTS,
    'passes_rights': settings.PASSES_RIGHTS,
}

def block_new(request):
    block = Block()
    
    form = BlockForm(
        request.POST or None,
        files=request.FILES or None,
    )
    if form.is_valid():
        form.save()  
        return redirect('blocks_all')
    return render(request, 'block_new.html', {'form': form, **rights,})


def block_edit(request, id):
    block = Block.objects.filter(id=id)[0]
    block.day = str(block.day)[:-15]
    form = BlockForm(
        request.POST or None,
        files=request.FILES or None,
        instance=block,
    )
    if form.is_valid():
        form.save()
        return redirect('blocks_all')
    return render(request, 'block_new.html', {'form': form, 'edit': True, **rights,})


def blocks_all(request):
    blocks = Block.objects.all()
    paginator = Paginator(blocks, 28)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)
    content = {'page': page,}
    return render(request, "pro_blocks_all.html", {**content, **rights,})


def blocks_sort(request, filtr):
    blocks = Block.objects.filter(block_type_id=filtr)
    paginator = Paginator(blocks, 28)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)
    content = {'page': page,}
    return render(request, "pro_blocks_all.html", {**content, **rights,})


def blocks_week(request, day, Print):
    #day = dt.datetime.now().date()
    day = dt.datetime.strptime(day, '%Y-%m-%d').date()
    
    dif = dt.datetime.now().date().isoweekday()
    start_day = day - dt.timedelta(days= dif - 1)
    end_day = day + dt.timedelta(days= 7 - dif)
    events = Event.objects.filter(day__range=[start_day, end_day], event_type_id=2)
    ids = []
    for event in events:
        ids.append(event.block_id)
    print(ids)
    blocks = Block.objects.all()

    f_lens = {'МФИ': {1: ''}, 'БМ': {2: ''}, 'ИВС': {3: ''}, 'МЦВ': {4: ''}, 'МК': {5: ''}, }
    for key, value in f_lens.items():
        for k, _ in value.items():
            start_year = dt.datetime.strptime(str(day.year) + '-01-01', '%Y-%m-%d').date()
            end_year = dt.datetime.strptime(str(day.year) + '-12-31', '%Y-%m-%d').date()
            year_count = Block.objects.filter(block_type_id=k, day__range=[start_year, end_year],).count()
            f_lens[key] = {k: str(year_count)}

    if Print == "True":
        Print = True
    else:
        Print = False
    content = {
        'blocks': blocks,
        'ids': ids,
        'f_lens': f_lens,  # 'МФИ': {id=1: year_count}
        'day': day,
        'url': settings.BASE_DIR + "\posts\static\pro\week_doc.pdf",
        'start_day': start_day,
        'end_day': end_day,
        'print': Print,
        }
    return render(request, "pro_blocks_week.html", {**content, **rights,})


def blocks(request, print):
    data = {'МФИ': 1, 'БМ': 2, 'ИВС': 3, 'МЦВ': 4, 'МК': 5, }
    dd = {'МФИ': [], 'БМ': [], 'ИВС': [], 'МЦВ': [], 'МК': [], }
    for key, value in data.items():
        on_injector = Block.objects.filter(block_type_id=value, block_state_id=2)
        on_storage = Block.objects.filter(block_type_id=value, block_state_id=1)
        dd[key].append({on_injector})
        dd[key].append({on_storage})
        max_len = max(len(on_injector), len(on_storage))
        in_data = []
        for i in range(max_len):
            para = []
            if i + 1 > len(on_injector):
                para.append('')
            else:
                para.append(on_injector[i])
            if i + 1 > len(on_storage):
                para.append('')
            else:
                para.append(on_storage[i])
            in_data.append(para)
            data[key] = in_data
    #print_stor_inj()
    if print == "True":
        print = True
    else:
        print = False
    content = {'data': data, 'dd': dd, 'url': settings.BASE_DIR + "\posts\static\pro\doc.pdf", 'print': print, }
    return render(request, 'pro_blocks.html', {**content, **rights,})


def on_where(request, block_id, on_where):
    block = Block.objects.filter(id=block_id)[0]
    if on_where == 'склад':
        block.block_state_id = 1
    else:
        block.block_state_id = 2
    block.save()
    return redirect('blocks', 'False')

def print_week(request, day):
    day = dt.datetime.strptime(day, '%Y-%m-%d').date()
    dif = dt.datetime.now().date().isoweekday()
    start_day = day - dt.timedelta(days= dif - 1)
    end_day = day + dt.timedelta(days= 7 - dif)
    
    start_year = dt.datetime.strptime(str(day.year) + '-01-01', '%Y-%m-%d').date()
    end_year = dt.datetime.strptime(str(day.year) + '-12-31', '%Y-%m-%d').date()

    mfi_year = [Block.objects.filter(block_type_id=1, day__range=[start_year, end_year],).count()]
    mfi_week = Block.objects.filter(block_type_id=1, day__range=[start_day, end_day], on_ready=True)
    mfi_on_control = Block.objects.filter(block_type_id=1, block_state_id=3, block_state_day__lte=end_day)
    mfi_on_fix = Block.objects.filter(block_type_id=1, block_state_id=4, block_state_day__lte=end_day)
    mfi_on_fix_developer = Block.objects.filter(block_type_id=1, block_state_id=5, block_state_day__lte=end_day)
    
    bm_year =  [Block.objects.filter(block_type_id=2, day__range=[start_year, end_year],).count()]
    bm_week = Block.objects.filter(block_type_id=2, day__range=[start_day, end_day], on_ready=True)
    bm_on_control = Block.objects.filter(block_type_id=2, block_state_id=3, block_state_day__lte=end_day)
    bm_on_fix = Block.objects.filter(block_type_id=2, block_state_id=4, block_state_day__lte=end_day)
    bm_on_fix_developer = Block.objects.filter(block_type_id=2, block_state_id=5, block_state_day__lte=end_day)
    
    ivs_year =  [Block.objects.filter(block_type_id=3, day__range=[start_year, end_year],).count()]
    ivs_week = Block.objects.filter(block_type_id=3, day__range=[start_day, end_day], on_ready=True)
    ivs_on_control = Block.objects.filter(block_type_id=3, block_state_id=3, block_state_day__lte=end_day)
    ivs_on_fix = Block.objects.filter(block_type_id=3, block_state_id=4, block_state_day__lte=end_day)
    ivs_on_fix_developer = Block.objects.filter(block_type_id=3, block_state_id=5, block_state_day__lte=end_day)
    
    mcv_year =  [Block.objects.filter(block_type_id=4, day__range=[start_year, end_year],).count()]
    mcv_week = Block.objects.filter(block_type_id=4, day__range=[start_day, end_day], on_ready=True)
    mcv_on_control = Block.objects.filter(block_type_id=4, block_state_id=3, block_state_day__lte=end_day)
    mcv_on_fix = Block.objects.filter(block_type_id=4, block_state_id=4, block_state_day__lte=end_day)
    mcv_on_fix_developer = Block.objects.filter(block_type_id=4, block_state_id=5, block_state_day__lte=end_day)
    
    mk_year =  [Block.objects.filter(block_type_id=5, day__range=[start_year, end_year],).count()]
    mk_week = Block.objects.filter(block_type_id=5, day__range=[start_day, end_day], on_ready=True)
    mk_on_control = Block.objects.filter(block_type_id=5, block_state_id=3, block_state_day__lte=end_day)
    mk_on_fix = Block.objects.filter(block_type_id=5, block_state_id=4, block_state_day__lte=end_day)
    mk_on_fix_developer = Block.objects.filter(block_type_id=5, block_state_id=5, block_state_day__lte=end_day)

    data = {
        'МФИ': {'A': mfi_year, 'B': mfi_week, 'C': mfi_on_control, 'D': mfi_on_fix, 'E': mfi_on_fix_developer, },
        'БМ': {'G': bm_year, 'H': bm_week, 'I': bm_on_control, 'J': bm_on_fix, 'k': bm_on_fix_developer},
        'ИВС': {'M': ivs_year, 'N': ivs_week, 'O': ivs_on_control, 'P': ivs_on_fix, 'Q': ivs_on_fix_developer},
        'МЦВ': {'S': mcv_year, 'T': mcv_week, 'U': mcv_on_control, 'V': mcv_on_fix, 'W': mcv_on_fix_developer},
        'МК': {'Y': mk_year, 'Z': mk_week, 'AA': mk_on_control, 'AB': mk_on_fix, 'AC': mk_on_fix_developer},

    }
    widths, full_len = get_width_for_title(data)
    start = '4'
    sample = settings.BASE_DIR + "\posts\static\pro\week_doc.xlsx"

    wb = openpyxl.load_workbook(sample)
    ws = wb['Производство']
    wb.remove(ws)
    wb.create_sheet(title='Производство')
    #ws = wb['Лист1']
    paint_all_borders(full_len, start, ['A', 'B', 'C', 'D', 'E', 'F'], wb['Производство'])
    write_data(
        data, widths, start, wb['Производство'],
        ['A2', 'A3', 'B2', 'B3', 'C2', 'C3', 'D2', 'E2', 'F2'],
        ['A3', 'B3', 'C3', 'D3', 'E3', 'F3'],
        ['A', 'B', 'C', 'D', 'E', 'F'],
    'Производство', start_day, end_day)
    '''
    for column_cells in wb['Склад-Изолятор'].columns:
        length = max(len(str(cell.value)) for cell in column_cells)/2
        wb['Склад-Изолятор'].column_dimensions[column_cells[0].column_letter].width = length
    '''
    wb['Производство'].column_dimensions['A'].width = 7
    wb['Производство'].column_dimensions['B'].width = 5
    wb['Производство'].column_dimensions['C'].width = 18
    wb['Производство'].column_dimensions['D'].width = 18
    wb['Производство'].column_dimensions['E'].width = 18
    wb['Производство'].column_dimensions['F'].width = 18
    wb.save(sample)

    #pythoncom.CoInitializeEx(0)
    #excel = client.Dispatch("Excel.Application")
    #excel.Visible = False
    # Read Excel File
    #sheets = excel.Workbooks.Open(sample)
    #work_sheets = sheets.Worksheets[0]
    # Convert into PDF File
    #pdf_file_name = (settings.BASE_DIR + "\posts\static\pro\week_doc.pdf")
    #work_sheets.ExportAsFixedFormat(0, pdf_file_name)
    #sheets.Close(True)
    return redirect('blocks_week', day, 'True')

def get_width_for_title(data):
    widths = []
    full_len = 0
    for key, value in data.items():
        max_len = 0
        for k, v in value.items():
            max_len = max(max_len, len(v))
        widths.append(max_len)
        full_len += max_len
    return widths, full_len


def print_stor_inj(request):
    mfi_on_injector = Block.objects.filter(block_type_id=1, block_state_id=2)
    mfi_on_storage = Block.objects.filter(block_type_id=1, block_state_id=1)
    bm_on_injector = Block.objects.filter(block_type_id=2, block_state_id=2)
    bm_on_storage = Block.objects.filter(block_type_id=2, block_state_id=1)
    ivs_on_injector = Block.objects.filter(block_type_id=3, block_state_id=2)
    ivs_on_storage = Block.objects.filter(block_type_id=3, block_state_id=1)
    mcv_on_injector = Block.objects.filter(block_type_id=4, block_state_id=2)
    mcv_on_storage = Block.objects.filter(block_type_id=4, block_state_id=1)
    mk_on_injector = Block.objects.filter(block_type_id=5, block_state_id=2)
    mk_on_storage = Block.objects.filter(block_type_id=5, block_state_id=1)

    data = {
        'МФИ': {'A': mfi_on_injector, 'B': mfi_on_storage, },
        'БМ': {'D': bm_on_injector, 'E': bm_on_storage, },
        'ИВС': {'G': ivs_on_injector, 'H': ivs_on_storage, },
        'МЦВ': {'J': mcv_on_injector, 'K': mcv_on_storage, },
        'МК': {'M': mk_on_injector, 'N': mk_on_storage, },
    }

    widths = []
    full_len = 0
    for _, value in data.items():
        max_len = 0
        for _, v in value.items():
            max_len = max(max_len, len(v))
        widths.append(max_len)
        full_len += max_len
    
    start = '2'
    sample = settings.BASE_DIR + "\posts\static\pro\doc.xlsx"

    wb = openpyxl.load_workbook(sample)
    ws = wb['Склад-Изолятор']
    wb.remove(ws)
    wb.create_sheet(title='Склад-Изолятор')
    #ws = wb['Лист1']
    paint_all_borders(full_len, start, ['A', 'B', 'C'], wb['Склад-Изолятор'])
    write_data(data, widths, start, wb['Склад-Изолятор'],
    ['A1', 'B1', 'C1'], ['A1', 'B1', 'C1'], ['A', 'B', 'C'],
    'Склад-Изолятор',)
    '''
    for column_cells in wb['Склад-Изолятор'].columns:
        length = max(len(str(cell.value)) for cell in column_cells)/2
        wb['Склад-Изолятор'].column_dimensions[column_cells[0].column_letter].width = length
    '''
    wb['Склад-Изолятор'].column_dimensions['A'].width = 10
    wb['Склад-Изолятор'].column_dimensions['B'].width = 38
    wb['Склад-Изолятор'].column_dimensions['C'].width = 38
    wb.save(sample)

    #pythoncom.CoInitializeEx(0)
    #excel = client.Dispatch("Excel.Application")
    #excel.Visible = False
    # Read Excel File
    #sheets = excel.Workbooks.Open(sample)
    #work_sheets = sheets.Worksheets[0]
    # Convert into PDF File
    #pdf_file_name = (settings.BASE_DIR + "\posts\static\pro\doc.pdf")
    #work_sheets.ExportAsFixedFormat(0, pdf_file_name)
    #sheets.Close(True)
    return redirect('blocks', 'True')

def write_data(data, widths, start, ws, mas1, mas2, mas3, P_S, start_day = 0, end_day= 0):
    thin = Side(border_style="thin", color="000000") # стиль границ
    medium = Side(border_style="medium", color="000000") # стиль границ
    if P_S == 'Производство':
        ws.merge_cells('A1:F1')
        start_day, end_day = str(start_day).split('-'), str(end_day).split('-')
        st, en = [], []
        for i in range(len(start_day)):
            st.append(start_day.pop())
        for i in range(len(end_day)):
            en.append(end_day.pop())
        start_day, end_day = '.'.join(st), '.'.join(en)
        ws['A1'] = 'Отчет за неделю (' + start_day + '-' + str(end_day) + ')'
        
        ws.merge_cells('B2:C2')
        ws['B2'] = 'Изготовлено'
        ws['B3'] = 'Год'
        ws['C3'] = 'Неделя'
        ws.merge_cells('D2:D3')
        ws['D2'] = 'Вх. Контроль'
        ws.merge_cells('E2:E3')
        ws['E2'] = 'Отремонтирован'
        ws.merge_cells('F2:F3')
        ws['F2'] = 'В ремонте у поставщика'
    else:
        ws['B1'] = 'Изолятор брака'
        ws['C1'] = 'Склад'
    for el in mas1:
        ws[el].fill = PatternFill("solid", "ABCDEF")
        ws[el].alignment = Alignment(horizontal="center", vertical="center")
        ws[el].border = Border(top=thin, left=thin, right=thin, bottom=thin,)
    ws['F2'].alignment = Alignment(wrap_text=True)
    for el in mas2:
        ws[el].border = Border(top=thin, left=thin, right=thin, bottom=medium,)

    for key, value in data.items():
        #print(key, value)
        width = widths.pop(0) - 1  # ширина    
        if width < 0:
            width = 0
        #print(width)
        start = 'A' + start
        end = start[:1] + str(int(start[1:]) + width)
        full = start + ':' + end   
        # Название блока
        ws.merge_cells(full)
        if P_S == 'Производство':
            start_b = 'B' + str(int(start.replace('A', 'B')[1:]))
            end = start_b[:1] + str(int(start_b[1:]) + width)
            full = start_b + ':' + end   
            ws.merge_cells(full)
        ws[start] = key
        ws[start].fill = PatternFill("solid", "ABCDEF")
        ws[start].alignment = Alignment(horizontal="center", vertical="center")
        symbl = ['B', 'C', 'D', 'E', 'F']
        for _, v in value.items():
            s = symbl.pop(0)
            i = 0

            for block in v:
                numbr = start.replace('A',s)[:1] + str(int(start.replace('A',s)[1:]) + i)
                try:
                    ws[numbr] = block.number + ' ' + block.info
                    ws[numbr].alignment = Alignment(wrap_text=True)
                except:
                    ws[start_b] = block
                    ws[start_b].alignment = Alignment(horizontal="center", vertical="center")
                    break
                i += 1
        for el in mas3:
            cell = el + str(int(start[1:]) + width)
            ws[cell].border = Border(top=thin, left=thin, right=thin, bottom=medium,)
        start = str(int(end[1:]) + 1)


def paint_all_borders(full_len, start, mass, ws):
    thin = Side(border_style="thin", color="000000") # стиль границ
    for i in range(int(start), full_len + int(start)):
        for el in mass:
            cell = el + str(i)
            ws[cell].border = Border(top=thin, left=thin, right=thin, bottom=thin,)


def pro_event_new(request, block_id):
    event = Event()
    event.block_id = block_id
    form = EventForm(
        request.POST or None,
        files=request.FILES or None,
        instance=event,
    )
    if form.is_valid():
        block = Block.objects.filter(id=block_id)[0]
        if event.event_type_id == 2:
            block.block_state_id = 4
            block.block_state_day = event.day
        elif event.event_type_id == 3:
            block.block_state_id = 1
            block.block_state_day = event.day
        if event.event_type_id == 4:
            block.block_state_id = 2
            block.block_state_day = event.day
        if event.event_type_id == 12:
            block.block_state_id = 3
            block.block_state_day = event.day
        if event.event_type_id == 13:
            block.block_state_id = 5
            block.block_state_day = event.day
        block.save()
        form.save()  
        return redirect('pro_events_all', block_id, )
    return render(request, 'event_new.html', {'form': form, **rights,})


def pro_events_all(request, block_id):
    number = Block.objects.filter(id=block_id)[0].number
    events = Event.objects.filter(block_id=block_id)
    paginator = Paginator(events, 28)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)
    content = {'page': page, 'number': number, }
    return render(request, "pro_events_all.html", {**content, **rights,})