from django.views.generic import CreateView
import re
import requests
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
User = get_user_model()
from .forms import CreationForm, User_infoForm, VacationForm, MessageForm
from django.core.paginator import Paginator
from django.conf import settings
from posts.models import Feedback
from django.shortcuts import render, get_object_or_404, redirect, reverse
from .models import User_info, Log, Vacation, User_widgets, Message
from django.db.models import Q
#from django.utils import timezone
import datetime as dt
from posts.models import Unit, Post
from tasks.models import Bord, Note, Task
from tasks.forms import NoteForm
from storage.models import User_units
from storage.models import Unit as Block

from calendar import monthrange
import openpyxl
from openpyxl.styles import  Border, Side, PatternFill, Font, Alignment
#import pythoncom
#from win32com import client
import json
import shutil
import os
from corresp.models import Corresp


holidays = [
    dt.datetime(2023, 1, 1).date(), dt.datetime(2023, 1, 2).date(),
    dt.datetime(2023, 1, 3).date(), dt.datetime(2023, 1, 4).date(),
    dt.datetime(2023, 1, 5).date(), dt.datetime(2023, 1, 6).date(),
    dt.datetime(2023, 1, 7).date(), dt.datetime(2023, 1, 8).date(),
    dt.datetime(2023, 2, 23).date(), dt.datetime(2023, 2, 24).date(),
    dt.datetime(2023, 3, 8).date(), dt.datetime(2023, 5, 1).date(),
    dt.datetime(2023, 5, 8).date(), dt.datetime(2023, 5, 9).date(),
    dt.datetime(2023, 6, 12).date(),
    dt.datetime(2023, 11, 6).date(), dt.datetime(2023, 12, 31).date(),
    dt.datetime(2024, 1, 1).date(), dt.datetime(2024, 1, 2).date(),
    dt.datetime(2024, 1, 3).date(), dt.datetime(2024, 1, 4).date(),
    dt.datetime(2024, 1, 5).date(), dt.datetime(2024, 1, 6).date(),
    dt.datetime(2024, 1, 7).date(), dt.datetime(2024, 1, 8).date(),
    dt.datetime(2024, 2, 23).date(), dt.datetime(2024, 3, 8).date(),
    dt.datetime(2024, 4, 29).date(), dt.datetime(2024, 4, 30).date(),
    dt.datetime(2024, 5, 1).date(), dt.datetime(2024, 5, 9).date(),
    dt.datetime(2024, 5, 10).date(), dt.datetime(2024, 6, 12).date(),
    dt.datetime(2024, 11, 4).date(), dt.datetime(2024, 12, 30).date(),
    dt.datetime(2024, 12, 31).date(),
]


def some_count(request):

    content = {
        'feedbacks_count': Feedback.objects.filter(state_id__in=[5, 6]).count, # Новое, В работе,
        'users_counts': User.objects.all().count,
        'users_with_pass_access_counts': User_info.objects.filter(pass_access=True).count,
        'users_with_reqs_access_counts': User_info.objects.filter(reqs_access=True).count,
        'users_with_stor_access_counts': User_info.objects.filter(stor_access=True).count,
        'users_with_task_access_counts': User_info.objects.filter(task_access=True).count,
        'users_with_user_access_counts': User_info.objects.filter(user_access=True).count,
        'users_with_corr_access_counts': User_info.objects.filter(corr_access=True).count,
        'users_with_conf_access_counts': User_info.objects.filter(conf_access=True).count,
        }
    if request.user.username in settings.RIGHTS:
        try:
            ips = {'10.1.98.247', '10.1.98.248', '10.1.98.249'} # не везде это нужно (tasks например)
            rele_count_of_1 = []
            for ip in ips:
                res = requests.get(
                    'http://admin:admin@' + ip + '/pstat.xml',
                    verify=False,
                    timeout=5,
                    )
                rele_count_of_1 += re.findall(r'[>][1][<]', str(res.content))
            content_plus = {'count_of_1': len(rele_count_of_1), }
        except:
            content_plus = {'count_of_1': 0, }
    return {**content, **content_plus}

def rights(request):
    rights = {
        'rights': settings.RIGHTS,
        'storage_rights': settings.STORAGE_RIGHTS,
        'sadec_rights': settings.SADEC_RIGHTS,
        'pro_rights': settings.PRO_RIGHTS,
        'passes_rights': settings.PASSES_RIGHTS,
        #**some_count(request), #если реле нет - тормозит работу сайта
        }
    return rights


class SignUp(CreateView):
    form_class = CreationForm
    success_url = reverse_lazy('login')
    template_name = "users/signup.html"


def users(request):
    users_list = User.objects.all().order_by('username')

    users_info = User_info.objects.all()
    paginator = Paginator(users_info, 40)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)

    content = {'users_info': users_info, }
    if request.user.username not in settings.RIGHTS:
        return redirect('index')
    add_log(
        request.user.id,
        dt.datetime.now(),
        "Переход на страницу",
        "Users",
        "Страница - Все сотрудники",
        'http://virtual2025.oak.cc:8000/' + 'auth/users/',
    )

    return render(request, 'users.html', {'page': page, **content, **rights(request)})


def users_access(request, app):

    if app == 'user':
        users_info = User_info.objects.filter(user_access=True)
    elif app == 'reqs':
        users_info = User_info.objects.filter(reqs_access=True)
    elif app == 'conf':
        users_info = User_info.objects.filter(conf_access=True)
    elif app == 'task':
        users_info = User_info.objects.filter(task_access=True)
    elif app == 'stor':
        users_info = User_info.objects.filter(stor_access=True)
    elif app == 'corr':
        users_info = User_info.objects.filter(corr_access=True)
    elif app == 'pass':
        users_info = User_info.objects.filter(pass_access=True)

    users_id = []
    for user in users_info:
        users_id.append(user.user_id)

    users = User.objects.filter(id__in=users_id).order_by('username')
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)

    content = {'users_info': users_info, }
    if request.user.username not in settings.RIGHTS:
        return redirect('index')
    add_log(
        request.user.id,
        dt.datetime.now(),
        "Переход на страницу",
        "Users",
        f"Страница - Права доступа({app})",
        'http://virtual2025.oak.cc:8000/' + f'auth/users_access/{app}/',
    )
    return render(request, 'users.html', {'page': page, **content, **rights(request)})


def users_in_otd(request, number):
    users_info = User_info.objects.filter(otd_number_id=number)
    users_id = []
    for user in users_info:
        users_id.append(user.user_id)

    users = User.objects.filter(id__in=users_id).order_by('username')
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)

    content = {'users_info': users_info, }
    if request.user.username not in settings.RIGHTS:
        return redirect('index')

    otd_number = Unit.objects.filter(id=number)[0].title
    if str(otd_number).rfind('(') != -1:
        otd_number = Unit.objects.filter(id=number)[0].title.split('(')[1][:-1]
    add_log(
        request.user.id,
        dt.datetime.now(),
        "Переход на страницу",
        "Users",
        f"Страница - Сотрудники отдела {otd_number}",
        'http://virtual2025.oak.cc:8000/' + f'auth/users_in_otd/{number}/',
    )
    return render(request, 'users.html', {'page': page, **content, **rights(request)})


@login_required
def users_info_change(request, user_id):
    info = User_info.objects.filter(user_id=user_id)[0]
    name = User.objects.filter(id=user_id)[0].get_full_name
    form = User_infoForm(
        request.POST or None,
        files=request.FILES or None,
        instance=info,
    )
    if form.is_valid():
        form.save()
        return redirect('user_space', user_id)
        # return redirect('users')
    content = {'form': form, 'edit': "edit", 'name': name, **rights(request), }
    return render(request, 'user_info_new.html', {**content, **rights(request), } )


def user_search(request):
    #users_info = User_info.objects.all()
    q = request.GET.get("q")
    users = User.objects.filter(
        Q(first_name__icontains=q) | Q(last_name__icontains=q)
    )
    users_id = [user.id for user in users]
    users_info = User_info.objects.filter(user_id__in=users_id)

    content = {'page': users_info,  }
    return render(request, 'users.html', {**content, **rights(request), })


def log_all(request):
    logs = Log.objects.all()
    paginator = Paginator(logs, 29)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)
    return render(request, 'log_all.html', {'page': page, **rights(request)},)


def add_log(*args):
    log = Log()
    log.user_id = args[0]
    log.day = args[1]
    log.event = args[2]
    log.res = args[3]
    log.before = args[4]
    log.after = args[5]
    log.save()


def del_vac_by_drop(request, otd, user_name, day):

    bosses = {
        'Николай Емельянов': [305, 306, 307],
        'Руслан Магомедов': [305, 306, 307],
    }

    year = day.split('-')[0]
    f_n = user_name.split(' ')[0]
    l_n = user_name.split(' ')[1]
    user = User.objects.filter(first_name=f_n, last_name=l_n)[0]

    if request.user.id != user.id:
        if request.user.get_full_name() not in bosses.keys():
            return redirect('vacations', year, otd)

    vacations = Vacation.objects.filter(user_id=user.id, year=str(year))
    date = dt.datetime.strptime(day, '%Y-%m-%d').date()
    for vac in vacations:
        if vac.day_start.date() <= date and vac.day_end.date() >= date:
            vac.delete()
            break
    return redirect('vacations', year, otd)


def add_new_vac(request, otd, day_s, month_s, year_s, long: int, day_e, month_e, user_id):
    date = dt.datetime.strptime(f'{year_s}-{month_s}-{day_s}', '%Y-%m-%d').date()
    date_end = dt.datetime.strptime(f'{year_s}-{month_e}-{day_e}', '%Y-%m-%d').date()
    vacation = Vacation()
    vacation.user_id = user_id
    vacation.day_start = date + dt.timedelta(days=1)

    print(date)

    i = 1
    if long != 0:
        while i < long:
            date += dt.timedelta(days=1)
            i += 1
            if date in holidays:
                i -= 1
        print(date)
        vacation.day_end = date + dt.timedelta(days=1)
        vacation.how_long = long
    else:
        while date < date_end:
            date += dt.timedelta(days=1)
            i += 1
            if date in holidays:
                i -= 1
        print(i)
        vacation.day_end = date_end + dt.timedelta(days=1)
        vacation.how_long = i
    vacation.year = year_s
    vacation.can_redact = True
    vacation.save()
    return redirect('vac_2', year_s, otd)


def get_key_from_dict_by_value(dict, value):
    return [k for k, v in dict.items() if v == value][0]


def get_cross_vacations(vacations, users_colors, month_num_str):
    data = []
    for i in range(len(vacations)):
        for j in range(i + 1, len(vacations)):
            if vacations[i] != vacations[j]:
                if (
                    (vacations[i].day_start >= vacations[j].day_start and vacations[i].day_end <= vacations[j].day_end)
                    or (vacations[i].day_start <= vacations[j].day_end and vacations[i].day_end > vacations[j].day_end)
                    or (vacations[i].day_start < vacations[j].day_start and vacations[i].day_end >= vacations[j].day_start)
                    or (vacations[i].day_start < vacations[j].day_start and vacations[i].day_end > vacations[j].day_end)
                ):
                    data.append(vacations[i])
                    data.append(vacations[j])
    vac_with_color = []
    for vac in set(data):
        d = {
            'vac': vac,
            'range': f"{vac.day_start.day} {month_num_str[vac.day_start.month][:3]} - {vac.day_end.day} {month_num_str[vac.day_end.month][:3]}",
            'color': users_colors[vac.user.get_full_name()],
        }
        vac_with_color.append(d)
    return vac_with_color


def copy_dict_for_js(month_all):
    month_all_for_js = {}
    for month, days in month_all.items():
        month_all_for_js[month] = {}
        for week, days_in_week in days.items():
            month_all_for_js[month][week] = {}
            for i in range(len(days_in_week)):
                month_all_for_js[month][week][i] = {}
                month_all_for_js[month][week][i]['name'] =month_all[month][week][i]['name']
                month_all_for_js[month][week][i]['date'] = {}
                for key in month_all[month][week][i]['data'].keys():
                    month_all_for_js[month][week][i]['date'][key.replace(' ', '_')] =  month_all[month][week][i]['data'][key]
                month_all_for_js[month][week][i]['data'] = month_all[month][week][i]['data']
    return month_all_for_js


def vac_2(request, year, otd):
    holidays = {
        2022: {
            'Январь': [1], 'Февраль': [23], 'Март': [8],
            'Апрель': [], 'Май': [1, 9], 'Июнь': [],
            'Июль': [], 'Август': [], 'Сентябрь': [],
            'Октябрь': [], 'Ноябрь': [4], 'Декабрь': [31],
        },
        2023: {
            'Январь': [1, 2, 3, 4, 5, 6, 7, 8], 'Февраль': [23, 24], 'Март': [8],
            'Апрель': [], 'Май': [1, 8, 9], 'Июнь': [12],
            'Июль': [], 'Август': [], 'Сентябрь': [],
            'Октябрь': [], 'Ноябрь': [6], 'Декабрь': [31],
        },
        2024: {
            'Январь': [1, 2, 3, 4, 5, 6, 7, 8], 'Февраль': [23], 'Март': [8],
            'Апрель': [29, 30], 'Май': [1, 9, 10], 'Июнь': [12],
            'Июль': [], 'Август': [], 'Сентябрь': [],
            'Октябрь': [], 'Ноябрь': [4], 'Декабрь': [30, 31],
        },

    }
    special_work_days = [
        dt.datetime(2024, 4, 27).date(),
        dt.datetime(2024, 11, 2).date(),
        dt.datetime(2024, 12, 28).date(),
    ]
    month_num_str = {
        1: 'Январь', 2: 'Февраль', 3: 'Март',
        4: 'Апрель', 5: 'Май', 6: 'Июнь',
        7: 'Июль', 8: 'Август', 9: 'Сентябрь',
        10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь',
    }
    some_colors = ["#B8860B", "#FF00FF", "#8A2BE2", "#00BFFF", "#F0E68C", "#7FFFD4", "#FFA500", "#20B2AA", "#FF96B4", "#00FA9A", "#FA8072",
        "#B8860B", "#FF00FF", "#8A2BE2", "#00BFFF", "#F0E68C", "#7FFFD4", "#FFA500", "#20B2AA", "#FF96B4", "#00FA9A", "#FA8072"]


    if year == 0:
        year = dt.datetime.today().year
    today = dt.datetime.today().date()
    month_all = full_year(year)


    bosses = {
        'Николай Емельянов': [305, 306, 307],
        'Руслан Магомедов': [305, 306, 307],
    }

    if request.user.get_full_name() not in bosses.keys():
        otd_id = User_info.objects.filter(user_id=request.user.id)[0].otd_number_id
        otd = int(Unit.objects.filter(id=otd_id)[0].description)
        otd_users = User_info.objects.filter(otd_number_id=otd_id)  # записи сотрудников отдела
        otd_users_id = [user.user_id for user in otd_users]  # id сотрудников отдела
        vacations = Vacation.objects.filter(user_id__in=otd_users_id, year=str(year))
        otds_for_choise = [otd]
        otd_users_full_names = [request.user]
    else:
        if otd == 0:  # Все
            vacations = Vacation.objects.filter(year=str(year))
        else:
            otd_id = Unit.objects.filter(description=otd)[0].id  # id отдела
            otd_users = User_info.objects.filter(otd_number_id=otd_id)  # записи сотрудников отдела
            otd_users_id = [user.user_id for user in otd_users]  # id сотрудников отдела
            vacations = Vacation.objects.filter(user_id__in=otd_users_id, year=str(year))
        otds_for_choise = bosses[request.user.get_full_name()]


        otd_ids = [Unit.objects.filter(description=descr)[0].id for descr in bosses[request.user.get_full_name()]]

        otd_users = User_info.objects.filter(otd_number_id__in=otd_ids)  # записи сотрудников отдела
        otd_users_id = [user.user_id for user in otd_users]
        otd_users_full_names = [user for user in User.objects.filter(id__in=otd_users_id)]




    users_colors = {}  # изменить на присвоение уникального номера юзеру

    for vac in vacations:
        if vac.user.get_full_name() not in users_colors.keys():
            users_colors[vac.user.get_full_name()] = some_colors.pop()

    # [{'vac': vac, 'range': range, 'color': color, }]
    cross_vacations = get_cross_vacations(vacations, users_colors, month_num_str)


    vacations_by_user = {}
    users_otd = User_info.objects.all()
    for vac in vacations:
        if vac.user.get_full_name() not in vacations_by_user.keys():
            vacations_by_user[vac.user.get_full_name()] = {
                'color': users_colors[vac.user.get_full_name()],
                'dates': [],  # {'d': 12.12 -18.12, 'vac_id': vac_id, 'vac_can_redact': vac_can_redact}
                'sum': 0,
                'otd': '',
                'user_id': vac.user_id,
            }
            for u in users_otd:
                if u.user_id == vac.user_id:
                    vacations_by_user[vac.user.get_full_name()]['otd'] =  u.otd_number
                    break
        vacations_by_user[vac.user.get_full_name()]['dates'].append(
            {'d': f"{vac.day_start.day} {month_num_str[vac.day_start.month][:3]} - {vac.day_end.day} {month_num_str[vac.day_end.month][:3]}",
            'vac_id': vac.id,
            'vac_can_redact': vac.can_redact,
            }
        )

        vacations_by_user[vac.user.get_full_name()]['sum'] += (vac.day_end.date() - vac.day_start.date()).days + 1

        for y, m_d in holidays.items():
            for m, d in m_d.items():
                for day in d:
                    month_number = get_key_from_dict_by_value(month_num_str, m)

                    date = today.replace(year=int(y), month=month_number, day=day)
                    if date >= vac.day_start.date() and date <= vac.day_end.date():
                        vacations_by_user[vac.user.get_full_name()]['sum'] -= 1

    for month, days in month_all.items():
        for week, days_in_week in days.items():
            for i in range(len(days_in_week)):
                data = {}
                date = ""
                if str(days_in_week[i]) != "":
                    month_number = get_key_from_dict_by_value(month_num_str, month)

                    day = today.replace(year=int(year), month=month_number, day=days_in_week[i])

                    for vac in vacations:
                        if day >= vac.day_start.date() and day <= vac.day_end.date():
                            data[vac.user.get_full_name()] = {
                                'date': f"{vac.day_start.date()} - {vac.day_end.date()}",
                                'color': users_colors[vac.user.get_full_name()],
                            }

                    m = [k for k, v in month_num_str.items() if v == month ][0]
                    date = today.replace(year=int(year), month=m, day=int(days_in_week[i]))
                month_all[month][week][i] = {'name': days_in_week[i], 'data': data, 'date': date, }

    month_all_for_js = copy_dict_for_js(month_all)

    
    
    if year in holidays.keys():
        h_days = holidays[year]
    else:
        h_days = {}

    all_vac_for_js = {}
    for vac in vacations:
        all_vac_for_js[vac.id] = [str(vac.day_start.date()), str(vac.day_end.date()), vac.how_long, vac.can_redact]

    
    return render(
        request,
        'vac_new_calendar.html',
        {**rights(request),
         'today': today,
         'year': year,
         'otd': otd,
         'pdf': settings.BASE_DIR + f"/posts/static/users/vacations_{otd}_{year}.pdf",
         **month_all,
         'json_data': json.dumps(month_all_for_js),
         'json_data_vacs': json.dumps(all_vac_for_js),
         
         'cross_vacations': cross_vacations,
         'len_cross_vacations': len(cross_vacations),
         'len_vacations': vacations.count,
         'special_work_days': special_work_days,
         'vacations_by_user': vacations_by_user,
         'holidays': h_days,
         'otds_for_choise': otds_for_choise,
         'bosses': [key for key in bosses.keys()],
         'otd_users_full_names': otd_users_full_names,
         }
    )


def vacations(request, year, otd):
    number = otd
    if otd == 0:
        return redirect('vacations_start')
    otd_id = Unit.objects.filter(description=otd)[0].id  # id отдела
    otd_users = User_info.objects.filter(otd_number_id=otd_id)  # записи сотрудников отдела
    otd_users_id = []
    #year = dt.datetime.today().year
    today = dt.datetime.today().date()
    for user in otd_users:
        otd_users_id.append(user.user_id)  # id сотрудников отдела
    vacations = Vacation.objects.filter(user_id__in=otd_users_id, year=str(year))

    months = {
        1:'Январь', 2:'Февраль', 3:'Март',
        4:'Апрель', 5:'Май', 6:'Июнь',
        7:'Июль', 8:'Август', 9:'Сентябрь',
        10:'Октябрь', 11:'Ноябрь', 12:'Декабрь',
    }

    full_year_by_month = {}
    vacations_by_month_by_users = {}

    for key in months.keys():
        full_year_by_month[months[key]] = []
        _, last_day = monthrange(year, key)
        for i in range(1,last_day + 1):
            day = str(year) + '-'
            
            if len(str(key)) == 1:
                day = day + '0' + str(key)
            else:
                day = day + str(key)
                
            if len(str(i)) == 1:
                day = day + '-0' + str(i)
            else:
                day = day + '-' + str(i)
            
            day = dt.datetime.strptime(day, '%Y-%m-%d').date()
            full_year_by_month[months[key]].append(day)

    for val in months.values():
        vacations_by_month_by_users[val] = {}
    for vacation in vacations:
        month_start = months[vacation.day_start.date().month]
        month_end = months[vacation.day_end.date().month]
        month_btw = []
        month_btw.append(month_start)
        if month_start not in vacations_by_month_by_users.keys():
            vacations_by_month_by_users[month_start] = {}
        if month_end not in vacations_by_month_by_users.keys():
            vacations_by_month_by_users[month_end] = {}

        for i in range(vacation.day_end.date().month - vacation.day_start.date().month):
            if (vacation.day_start.date().month + i) > vacation.day_start.date().month and (vacation.day_start.date().month + i) < vacation.day_end.date().month:
                m = months[vacation.day_start.date().month + i]

                if m not in vacations_by_month_by_users.keys():
                    vacations_by_month_by_users[m] = {}
                month_btw.append(m)
        if month_start != month_end:
            month_btw.append(month_end)
        day_start = vacation.day_start.date()
        day_end = vacation.day_end.date()
        name = vacation.user.get_full_name()

        for m in month_btw:
            if name not in vacations_by_month_by_users[m].keys():
                vacations_by_month_by_users[m][name] = []

            vacations_by_month_by_users[m][name].append(day_start)
            while day_start < day_end:
                some_day = day_start + dt.timedelta(days=1)
                vacations_by_month_by_users[m][name].append(some_day)
                day_start = some_day
            day_start = vacation.day_start.date()           

    for k, v  in vacations_by_month_by_users.items():
        vacations_by_month_by_users[k] = dict(sorted(v.items()))

    #for_js = {} # {"20230311": [id1, id2...]}
    #users_js = {} # {"20230311": [Емельянов Николай, id2...]}
    #colors_js = {} # {"20230311": [#FF6384, #63FF84...]}
    #user_ids = {} # {"id": [Емельянов Николай, ...]}

    some_colors = ["#B8860B", "#FF00FF", "#8A2BE2", "#00BFFF", "#F0E68C", "#7FFFD4", "#FFA500", "#20B2AA", "#FF96B4", "#00FA9A", "#FA8072"]
    user_colors = {}
    for vacation in vacations:
        if vacation.user.get_full_name() not in user_colors.keys():
            user_colors[vacation.user.get_full_name()] = some_colors.pop()

    days_can_redact = {}  # словарь типа "День:[Человек_возможность редактировать,]"
    for vacation in vacations:
        some_day = vacation.day_start
        while some_day <= vacation.day_end:
            if str(some_day.date()).replace('-', '') not in days_can_redact.keys():
                days_can_redact[str(some_day.date()).replace('-', '')] = []
            if str(vacation.user.get_full_name()) not in  days_can_redact[str(some_day.date()).replace('-', '')]:
                days_can_redact[str(some_day.date()).replace('-', '')].append(str(vacation.user.get_full_name()) + "_" + str(vacation.can_redact))
            some_day = some_day  + dt.timedelta(days=1)
    print(days_can_redact)
    months_to_numbers = {'января': '01', 'февраля': '02', 'марта': '03', 'апреля': '04',
                         'мая': '05', 'июня': '06', 'июля': '07', 'августа': '08',
                         'сентября': '09', 'октября': '10', 'ноября': '11', 'декабря': '12',}
    years = [2023, 2024]
    bosses = {
        'Николай Емельянов': [305, 306, 307],
        'Руслан Магомедов': [305, 306, 307],
    }
    try:
        otd_id = User_info.objects.filter(user_id=request.user.id)[0].otd_number_id
        otd = int(Unit.objects.filter(id=otd_id)[0].description)
    except:
        otd = 0
    

    holidays = [
        dt.datetime(2023,1,1).date(), dt.datetime(2023,1,2).date(), dt.datetime(2023,1,3).date(),
        dt.datetime(2023,1,4).date(), dt.datetime(2023,1,5).date(), dt.datetime(2023,1,6).date(),
        dt.datetime(2023,1,7).date(), dt.datetime(2023,1,8).date(), dt.datetime(2023,2,23).date(),
        dt.datetime(2023,2,24).date(), dt.datetime(2023,3,8).date(), dt.datetime(2023,5,1).date(),
        dt.datetime(2023,5,8).date(), dt.datetime(2023,5,9).date(), dt.datetime(2023,6,12).date(),
        dt.datetime(2023,11,6).date(), dt.datetime(2023,12,31).date(),
        dt.datetime(2024, 1, 1).date(), dt.datetime(2024, 1, 2).date(),
        dt.datetime(2024, 1, 3).date(), dt.datetime(2024, 1, 4).date(),
        dt.datetime(2024, 1, 5).date(), dt.datetime(2024, 1, 6).date(),
        dt.datetime(2024, 1, 7).date(), dt.datetime(2024, 1, 8).date(),
        dt.datetime(2024, 2, 23).date(), dt.datetime(2024, 3, 8).date(),
        dt.datetime(2024, 4, 29).date(), dt.datetime(2024, 4, 30).date(),
        dt.datetime(2024, 5, 1).date(), dt.datetime(2024, 5, 9).date(),
        dt.datetime(2024, 5, 10).date(), dt.datetime(2024, 6, 12).date(),
        dt.datetime(2024, 11, 4).date(), dt.datetime(2024, 12, 30).date(),
        dt.datetime(2024, 12, 31).date(),
    ]
    special_work_days = [
        dt.datetime(2024, 4, 27).date(),
        dt.datetime(2024, 11, 2).date(),
        dt.datetime(2024, 12, 28).date(),
    ]
    return render(
        request,
        'vacations_all.html',
        {
            **rights(request),
            'pdf': settings.BASE_DIR + f"/posts/static/users/vacations_{number}_{year}.pdf",
            'holidays': holidays,
            'special_work_days': special_work_days,
            'year': year, 'years': years,
            'json_data': json.dumps(days_can_redact),
            'months_to_numbers': json.dumps(months_to_numbers),
            'bosses': bosses,
            'today': today,
            'vacations': vacations,
            'otd': otd, 'number': number,
            'user_colors': user_colors,
            'full_year_by_month': full_year_by_month,
            'vacations_by_month_by_users': vacations_by_month_by_users,
        },
    )


def get_user_bords_and_ids(all_bords, user_id):
    user_bords = [b for b in all_bords if (str(user_id) in b.guests.split('_')) or b.user_id == user_id ]
    user_bords_ids = [bord.id for bord in user_bords]
    if len(user_bords_ids) == 0:
        user_bords_ids.append(0)
    return user_bords, user_bords_ids

def get_tasks_sorted_by_bords(user_bords_ids):
    data = {}
    all_tasks = Task.objects.all()
    for t in all_tasks:
        if t.bord_id in user_bords_ids:
            if t.bord_id not in data.keys():
                data[t.bord_id] = []
            data[t.bord_id].append(t)
    return data

def get_new_tasks_sorted_by_bords(tasks_sorted_by_bords, user):
    data = {}
    for key in tasks_sorted_by_bords.keys():
        for t in tasks_sorted_by_bords[key]:
            if t.slave == user:
                if t.new:
                    if key not in data.keys():
                        data[key] = 0
                    data[key] += 1
    return data

@login_required
def user_space(request, user_id):
    if user_id == 0:
        return redirect('user_space', request.user.id)
    users = User.objects.all()
    users_info = User_info.objects.all()
    users_info_for_widget = {}
    for u in users_info:
        if u.otd_number not in users_info_for_widget.keys():
            users_info_for_widget[u.otd_number] = {}
        users_info_for_widget[u.otd_number][u.user_id] = get_object_or_404(User, id=u.user_id)

    user = get_object_or_404(User, id=user_id)
    user_info = get_object_or_404(User_info, user_id=user_id)
    user_widgets = get_object_or_404(User_widgets, user_id=user_id)
    widgets_order = str(user_widgets.widgets_order)
    # Работа с досками задач -------------------------------------------------
    # Все доски с задачами
    all_bords = Bord.objects.all()
    # Все доски [], где участвует пользователь и их id []

    user_bords, user_bords_ids = get_user_bords_and_ids(all_bords, user_id)
    tasks_sorted_by_bords = get_tasks_sorted_by_bords(user_bords_ids)
    new_tasks_sorted_by_bords = get_new_tasks_sorted_by_bords(tasks_sorted_by_bords, request.user)
    # ------------------------------------------------------------------------

    posts = Post.objects.filter(author_id=user.id).order_by('-id')[:10]
    vacations = Vacation.objects.filter(user_id=user.id, year=str(dt.datetime.today().year))
    chats = Message.objects.filter(
        Q(user_one_id=user_id) | Q(user_two_id=user_id)
    )
    user_chats = set()
    for c in chats:
        user_chats.add(User.objects.filter(id=c.user_one_id)[0])
        user_chats.add(User.objects.filter(id=c.user_two_id)[0])

    unreaded = {}
    for u in user_chats:
        for chat in chats:
            if (u.id == chat.user_two_id or u.id == chat.user_one_id) and u.id != chat.witch_write_id and not chat.readed:
                if u.id not in unreaded.keys():
                    if u.id == chat.user_two_id:
                        unreaded[chat.user_one_id] = 0
                    else:
                        unreaded[chat.user_two_id] = 0     
                try:                 
                    unreaded[chat.user_one_id] += 1
                except:
                    unreaded[chat.user_two_id] += 1

    user_notes = Note.objects.filter(user=request.user)
    json_data ={}
    for t in user_notes:

        json_data[t.id] = []
        json_data[t.id].append(t.text)
        json_data[t.id].append(t.same_id)

        if t.day_update is not None:
            json_data[t.id].append(str(t.day_update.date()))
        else:
            json_data[t.id].append("")

    note = Note()
    note.user = request.user
    form = NoteForm(
        request.POST or None,
        files=request.FILES or None,
        instance=note,
    )
    if form.is_valid():
        note = form.save(commit=False)
        if note.same_id is not None:
            note.id = int(note.same_id)
        else: 
            try:
                note.id = Note.objects.all().latest('id').id + 1
            except:
                note.id = 1
            note.same_id = note.id
        note.save()
        return redirect('user_space', request.user.id)

    #print(unreaded)
    today = dt.datetime.today().date
    #  вставка из vac_2
    holidays = {
        2024: {
            'Январь': [1, 2, 3, 4, 5, 6, 7, 8], 'Февраль': [23], 'Март': [8],
            'Апрель': [29, 30], 'Май': [1, 9, 10], 'Июнь': [12],
            'Июль': [], 'Август': [], 'Сентябрь': [],
            'Октябрь': [], 'Ноябрь': [4], 'Декабрь': [30, 31],
        },
    }
    special_work_days = [
        dt.datetime(2024, 4, 27).date(),
        dt.datetime(2024, 11, 2).date(),
        dt.datetime(2024, 12, 28).date(),
    ]
    month_num_str = {
        1: 'Январь', 2: 'Февраль', 3: 'Март',
        4: 'Апрель', 5: 'Май', 6: 'Июнь',
        7: 'Июль', 8: 'Август', 9: 'Сентябрь',
        10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь',
    }
    some_colors = ["#FA8072"]
    year = dt.datetime.today().year
    today = dt.datetime.today().date()
    month_all = full_year(year)

    users_colors = {}  # изменить на присвоение уникального номера юзеру

    for vac in vacations:
        if vac.user.get_full_name() not in users_colors.keys():
            users_colors[vac.user.get_full_name()] = some_colors.pop()

    # [{'vac': vac, 'range': range, 'color': color, }]
    cross_vacations = get_cross_vacations(vacations, users_colors, month_num_str)


    vacations_by_user = {}
    users_otd = User_info.objects.all()
    for vac in vacations:
        if vac.user.get_full_name() not in vacations_by_user.keys():
            vacations_by_user[vac.user.get_full_name()] = {
                'color': users_colors[vac.user.get_full_name()],
                'dates': [],  # {'d': 12.12 -18.12, 'vac_id': vac_id, 'vac_can_redact': vac_can_redact}
                'sum': 0,
                'otd': '',
                'user_id': vac.user_id,
            }
            for u in users_otd:
                if u.user_id == vac.user_id:
                    vacations_by_user[vac.user.get_full_name()]['otd'] =  u.otd_number
                    break
        vacations_by_user[vac.user.get_full_name()]['dates'].append(
            {'d': f"{vac.day_start.day} {month_num_str[vac.day_start.month][:3]} - {vac.day_end.day} {month_num_str[vac.day_end.month][:3]}",
            'vac_id': vac.id,
            'vac_can_redact': vac.can_redact,
            }
        )

        vacations_by_user[vac.user.get_full_name()]['sum'] += (vac.day_end.date() - vac.day_start.date()).days + 1

        for y, m_d in holidays.items():
            for m, d in m_d.items():
                for day in d:
                    month_number = get_key_from_dict_by_value(month_num_str, m)

                    date = today.replace(year=int(y), month=month_number, day=day)
                    if date >= vac.day_start.date() and date <= vac.day_end.date():
                        vacations_by_user[vac.user.get_full_name()]['sum'] -= 1

    for month, days in month_all.items():
        for week, days_in_week in days.items():
            for i in range(len(days_in_week)):
                data = {}
                date = ""
                if str(days_in_week[i]) != "":
                    month_number = get_key_from_dict_by_value(month_num_str, month)

                    day = today.replace(year=int(year), month=month_number, day=days_in_week[i])

                    for vac in vacations:
                        if day >= vac.day_start.date() and day <= vac.day_end.date():
                            data[vac.user.get_full_name()] = {
                                'date': f"{vac.day_start.date()} - {vac.day_end.date()}",
                                'color': users_colors[vac.user.get_full_name()],
                            }

                    m = [k for k, v in month_num_str.items() if v == month ][0]
                    date = today.replace(year=int(year), month=m, day=int(days_in_week[i]))
                month_all[month][week][i] = {'name': days_in_week[i], 'data': data, 'date': date, }
    if year in holidays.keys():
        h_days = holidays[year]
    else:
        h_days = {}

    storage_user_likes = User_units.objects.filter(user_id=request.user.id)
    storage_user_likes_ids = [like.unit_id for like in storage_user_likes]
    favorite_units = Block.objects.filter(id__in=storage_user_likes_ids)

    corresp = Corresp.objects.all()
    cor_years = [c.day.year for c in corresp]
    cor_years = sorted(set(cor_years))
    cor_tags = [str(c.tag) for c in corresp]
    cor_tags = sorted(set(cor_tags))

    return render(
        request,
        'user_space.html', {
            'today': today,
            'user': user,
            'users': users,
            'user_info': user_info,
            'users_info': users_info,
            'user_widgets': user_widgets,
            'widgets_order': widgets_order,
            'user_notes': user_notes,
            'user_bords': user_bords,
            'user_start_bord_id': user_bords_ids[0],
            'posts': posts,
            'vacations': vacations,
            'user_chats': user_chats,
            'unreaded': unreaded,
            'form': form,
            'json_data': json.dumps(json_data),
            'year': year,
            **month_all,
            'cross_vacations': cross_vacations,
            'len_cross_vacations': len(cross_vacations),
            'len_vacations': vacations.count,
            'special_work_days': special_work_days,
            'vacations_by_user': vacations_by_user,
            'holidays': h_days,
            'favorite_units': favorite_units,
            'users_info_for_widget': users_info_for_widget,
            'new_tasks_sorted_by_bords': new_tasks_sorted_by_bords,
            'corresp': corresp,
            'cor_years': cor_years,
            'cor_tags': cor_tags,
        }
    )


@login_required
def backup_base(request):
    if os.path.exists(settings.MEDIA_ROOT + f"/db.sqlite3"):
        os.remove(settings.MEDIA_ROOT + f"/db.sqlite3")
    shutil.copy2(settings.BASE_DIR + f"/db.sqlite3", settings.MEDIA_ROOT + f"/db.sqlite3")
    return redirect(request.META.get('HTTP_REFERER'))


@login_required
def user_widget_add(request, user_id, widget_name):
    user_info = get_object_or_404(User_info, user_id=user_id)
    user_widgets = get_object_or_404(User_widgets, user_id=user_id)

    names = {
        'reqs': 'reqs_access',
        'vacs': 'vacs_access',
        'test': 'test_access',
        'corr': 'corr_access',
        'task': 'task_access',
        'bibl': 'bibl_access',
        'stor': 'stor_access',
        'users': 'user_access',
        'news': 'news_access',
        'mess': 'mess_access',
        'notes': 'note_access',
        'calc': 'calc_access',
    }
    if names[widget_name] == 'reqs_access':
        if user_info.reqs_access and not user_widgets.reqs:
            user_widgets.reqs = True
    elif names[widget_name] == 'vacs_access':
        if user_info.vacs_access and not user_widgets.vacs:
            user_widgets.vacs = True
    elif names[widget_name] == 'test_access':
        if user_info.test_access and not user_widgets.test:
            user_widgets.test = True
    elif names[widget_name] == 'corr_access':
        if user_info.corr_access and not user_widgets.corr:
            user_widgets.corr = True
    elif names[widget_name] == 'task_access':
        if user_info.task_access and not user_widgets.task:
            user_widgets.task = True
    elif names[widget_name] == 'bibl_access':
        if user_info.bibl_access and not user_widgets.bibl:
            user_widgets.bibl = True
    elif names[widget_name] == 'stor_access':
        if user_info.stor_access and not user_widgets.stor:
            user_widgets.stor = True
    elif names[widget_name] == 'user_access':
        if user_info.user_access and not user_widgets.users:
            user_widgets.users = True
    elif names[widget_name] == 'news_access':
        if user_info.news_access and not user_widgets.news:
            user_widgets.news = True
    elif names[widget_name] == 'mess_access':
        if user_info.mess_access and not user_widgets.mess:
            user_widgets.mess = True
    elif names[widget_name] == 'note_access':
        if not user_widgets.notes:
            user_widgets.notes = True
    elif names[widget_name] == 'calc_access':
        if not user_widgets.calc:
            user_widgets.calc = True
    user_widgets.save()

    return redirect('user_space', user_id)


@login_required
def user_widget_close(request, user_id, widget_name,  widget_id):
    user_widgets = get_object_or_404(User_widgets, user_id=user_id)

    names = {
        'reqs': 'reqs_open',
        'vacs': 'vacs_open',
        'test': 'test_open',
        'corr': 'corr_open',
        'tasks': 'task_open',
        'bibl': 'bibl_open',
        'stor': 'stor_open',
        'users': 'users_open',
        'news': 'news_open',
        'mess': 'mess_open',
        'notes': 'notes_open',
        'calc': 'calc_open',
    }


    order = getattr(user_widgets, 'widgets_order').split('_')
    if getattr(user_widgets, names[widget_name]):
        setattr(user_widgets, names[widget_name], False)

        order.append(order.pop(order.index(str(widget_id))))
        setattr(user_widgets, 'widgets_order', '_'.join(order))
    else:
        setattr(user_widgets, names[widget_name], True)

        order.insert(0, order.pop(order.index(str(widget_id))))
        setattr(user_widgets, 'widgets_order', '_'.join(order))


            
    user_widgets.save()
    return redirect(request.META.get('HTTP_REFERER'))
    #return redirect('user_space', user_id)


@login_required
def user_widget_close_all(request, user_id, state):
    user_widgets = get_object_or_404(User_widgets, user_id=user_id)

    names = {
        'reqs': 'reqs_open',
        'vacs': 'vacs_open',
        'test': 'test_open',
        'corr': 'corr_open',
        'tasks': 'task_open',
        'bibl': 'bibl_open',
        'stor': 'stor_open',
        'users': 'users_open',
        'news': 'news_open',
        'mess': 'mess_open',
        'notes': 'notes_open',
        'calc': 'calc_open',
    }
    if state == 'close':
        for name in names.keys():
            setattr(user_widgets, names[name], False)
    elif state == 'open':
        for name in names.keys():
            setattr(user_widgets, names[name], True)

            
    user_widgets.save()
    return redirect(request.META.get('HTTP_REFERER'))


@login_required
def user_widget_delete(request, user_id, widget_name):
    user_info = get_object_or_404(User_info, user_id=user_id)
    user_widgets = get_object_or_404(User_widgets, user_id=user_id)

    names = {
        'reqs': 'reqs_access',
        'vacs': 'vacs_access',
        'test': 'test_access',
        'corr': 'corr_access',
        'task': 'task_access',
        'bibl': 'bibl_access',
        'stor': 'stor_access',
        'users': 'user_access',
        'news': 'news_access',
        'mess': 'mess_access',
        'notes': 'note_access',
        'calc': 'calc_access',
    }
    if names[widget_name] == 'reqs_access':
        if user_widgets.reqs:
            user_widgets.reqs = False
    elif names[widget_name] == 'vacs_access':
        if user_widgets.vacs:
            user_widgets.vacs = False
    elif names[widget_name] == 'test_access':
        if user_widgets.test:
            user_widgets.test = False
    elif names[widget_name] == 'corr_access':
        if user_widgets.corr:
            user_widgets.corr = False
    elif names[widget_name] == 'task_access':
        if user_widgets.task:
            user_widgets.task = False
    elif names[widget_name] == 'bibl_access':
        if user_widgets.bibl:
            user_widgets.bibl = False
    elif names[widget_name] == 'stor_access':
        if user_widgets.stor:
            user_widgets.stor = False
    elif names[widget_name] == 'user_access':
        if user_widgets.users:
            user_widgets.users = False
    elif names[widget_name] == 'news_access':
        if user_widgets.news:
            user_widgets.news = False
    elif names[widget_name] == 'mess_access':
        if user_widgets.mess:
            user_widgets.mess = False
    elif names[widget_name] == 'note_access':
        if user_widgets.notes:
            user_widgets.notes = False
    elif names[widget_name] == 'calc_access':
        if user_widgets.calc:
            user_widgets.calc = False
    user_widgets.save()

    return redirect('user_space', user_id)


@login_required
def note_delete(request, note_id):
    note = get_object_or_404(Note, id=note_id)
    note.delete()
    return redirect('user_space', request.user.id)

@login_required
def messages(request, user_one_id, user_two_id):
    user_one = get_object_or_404(User, id=user_one_id)
    user_two = get_object_or_404(User, id=user_two_id)
    chats = Message.objects.filter(
        Q(user_one_id=request.user.id) | Q(user_two_id=request.user.id)
    )
    user_chats = set()
    for c in chats:
        user_chats.add(User.objects.filter(id=c.user_one_id)[0])
        user_chats.add(User.objects.filter(id=c.user_two_id)[0])
    
    unreaded = {}
    for user in user_chats:
        for chat in chats:
            if (user.id == chat.user_two_id or user.id == chat.user_one_id) and user.id != chat.witch_write_id and not chat.readed:
                if user.id not in unreaded.keys():
                    if user.id == chat.user_two_id:
                        unreaded[chat.user_one_id] = 0
                    else:
                        unreaded[chat.user_two_id] = 0
                try:
                    unreaded[chat.user_one_id] += 1
                    
                except:
                    unreaded[chat.user_two_id] += 1

    messages = Message.objects.filter(
        Q(user_one_id=user_one_id, user_two_id=user_two_id) | Q(user_one_id=user_two_id, user_two_id=user_one_id)
    )
    for m in messages:
        if m.witch_write_id != request.user.id:
            
            if m.readed == False:
                m.readed = True
                m.save()

    mes = Message()
    mes.user_one_id = user_one_id
    mes.user_two_id = user_two_id
    mes.witch_write_id = request.user.id
    mes.pub_date = dt.datetime.now()
    mes.user_one_id = user_one_id
    form = MessageForm(
        request.POST or None,
        files=request.FILES or None,
        instance=mes,
    )
    if form.is_valid():
        mes = form.save(commit=False)
        if mes.text != "":
            mes.save()
            return redirect('messages', user_one_id, user_two_id)
            
    return render(
        request,
        'user_messages.html',
        {'user': request.user, **rights(request), 'messages': messages, 'user_one': user_one, 'user_two': user_two, 'form': form,
        'user_chats': user_chats, 'unreaded': unreaded, 'users': User.objects.all()}
    )


@login_required
def vacations_by_user(request, year, otd):
    vacations = Vacation.objects.filter(user_id=request.user.id, year=str(year))
    bosses = {'Николай Емельянов': [305, 306, 307], 'Руслан Магомедов': [305, 306, 307], },
    years = [2023, 2024]
    return render(
        request,
        'vacations_by_user.html',
        {'user': request.user, **rights(request), 'otd': otd, 'number': otd,  'vacations': vacations, 'year': year, 'years': years, 'bosses': bosses,
        'pdf': settings.BASE_DIR + f"/posts/static/users/vacations_{otd}.pdf",}
    )

def vacation_new(request, year, otd):
    vacation = Vacation()
    vacation.user_id = request.user.id
    form = VacationForm(
        request.POST or None,
        files=request.FILES or None,
        instance=vacation,
    )
    if form.is_valid():
        vac = form.save(commit=False)
        if vac.day_start is None or vac.day_end is None or vac.day_start >= vac.day_end:
            return render(
                request,
                'vacation_new.html',
                {'form': form, 'user': request.user, **rights(request), 'year': year, 'otd': otd, }
            )
        vac.save()
        return redirect('vac_2', year, otd)
    return render(
        request,
        'vacation_new.html',
        {'form': form, 'user': request.user, **rights(request), 'otd': otd, 'number': otd, 'year': year,
        'bosses':{'Николай Емельянов': [305, 306, 307], 'Руслан Магомедов': [305, 306, 307], },
        'pdf': settings.BASE_DIR + f"/posts/static/users/vacations_{otd}_{year}.pdf", }
    )

@login_required

def vac_edit(request, otd, day_s, month_s, year_s, long: int, day_e, month_e, user_id, vac_id):
    date = dt.datetime.strptime(f'{year_s}-{month_s}-{day_s}', '%Y-%m-%d').date()
    date_end = dt.datetime.strptime(f'{year_s}-{month_e}-{day_e}', '%Y-%m-%d').date()

    vacation = get_object_or_404(Vacation, id=vac_id)
    vacation.user_id = user_id
    vacation.day_start = date + dt.timedelta(days=1)

    print(date)

    i = 1
    if long != 0:
        while i < long:
            date += dt.timedelta(days=1)
            i += 1
            if date in holidays:
                i -= 1
        print(date)
        vacation.day_end = date + dt.timedelta(days=1)
        vacation.how_long = long
    else:
        while date < date_end:
            date += dt.timedelta(days=1)
            i += 1
            if date in holidays:
                i -= 1
        print(i)
        vacation.day_end = date_end + dt.timedelta(days=1)
        vacation.how_long = i
    vacation.year = year_s
    vacation.can_redact = True
    vacation.save()
    return redirect('vac_2', year_s, otd)

def vacation_edit(request, year, otd, vac_id):
    bosses = {
        'Николай Емельянов': [305, 306, 307],
        'Руслан Магомедов': [305, 306, 307],
    }

    vac = get_object_or_404(Vacation, id=vac_id)
    vac.day_start = str(vac.day_start)[:-15]
    vac.day_end = str(vac.day_end)[:-15]
    form = VacationForm(request.POST or None,
                    files=request.FILES or None,
                    instance=vac)
    if form.is_valid():
        vac = form.save(commit=False)
        if vac.day_start >= vac.day_end:
            return render(
                request,
                'vacation_new.html',
                {'form': form, 'user': vac.user, **rights(request), 'vac_id': vac_id, 'otd': otd, }
            )
        vac.save()
        return redirect('vac_2', year, otd)
    return render(
        request,
        'vacation_new.html',
        {
            'form': form,
            'user': request.user,
            'edit': True,
            **rights(request),
            'vac_id': vac_id,
            'otd': otd,
            'number': otd,
            'year': year,
            'bosses': [key for key in bosses.keys()],
        }
    )

@login_required
def vacation_confirm(request, year, otd, vac_id):
    vac = get_object_or_404(Vacation, id=vac_id)
    if vac.can_redact:
        vac.can_redact = False
    else:
        vac.can_redact = True
    vac.save()
    return redirect('vac_2', year, otd )

@login_required
def vacation_confirm_from_day(request, year, otd, user, day):
    day = dt.datetime.strptime(day, '%Y-%m-%d')
    for u in User.objects.all():
        if u.get_full_name() == user:
            print(u.id)
            vac = get_object_or_404(Vacation, user_id=u.id, day_start__lte=day, day_end__gte=day)
            if vac.can_redact:
                vac.can_redact = False
            else:
                vac.can_redact = True
            vac.save()
            break


    return redirect('vacations', year, otd)

@login_required
def vacation_delete(request, otd, year, vac_id):
    vac = get_object_or_404(Vacation, id=vac_id)
    vac.delete()
    return redirect('vac_2', year, otd)


def paint_all_borders(full_len, start, mass, ws):
    thin = Side(border_style="thin", color="000000") # стиль границ
    for i in range(int(start), full_len + int(start)):
        for el in mass:
            cell = el + str(i)
            ws[cell].border = Border(top=thin, left=thin, right=thin, bottom=thin,)

'''
def vacations_print(request, year, otd):
    #year = str(dt.datetime.now().year)
    if otd == 0:
        vacations = Vacation.objects.filter(year=year)
    else:
        otd_id = Unit.objects.filter(description=otd)[0].id  # id отдела
        otd_users = User_info.objects.filter(otd_number_id=otd_id)  # записи сотрудников отдела
        otd_users_id = []
        for user in otd_users:
            otd_users_id.append(user.user_id)  # id сотрудников отдела
        vacations = Vacation.objects.filter(user_id__in=otd_users_id, year=year)

    data = {}
    
    for vac in vacations:
        user = User.objects.filter(id=vac.user_id)[0]
        if str(user.get_full_name()) not in data.keys():
            data[str(user.get_full_name())] = []
        data[str(user.get_full_name())].append(vac)
  
    start = '2'
    sample = settings.BASE_DIR + '/posts/static/users/vacations.xlsx'
    try:
        wb = openpyxl.load_workbook(sample)
        ws = wb['Лист1']
        widths = []
        wb.remove(ws)
        wb.create_sheet(title='Лист1')
        ws = wb['Лист1']
        full_len = 0

        for _, value in data.items():
            
            max_len = len(value)
            widths.append(max_len)
            full_len += max_len
        paint_all_borders(full_len, 2, ['A', 'B', 'C', 'D', 'E'], wb['Лист1'])
        ws['A1'] = 'Сотрудник'
        ws['B1'] = 'Дата начала'
        ws['C1'] = 'Дата окончания'
        ws['D1'] = 'Длительность'
        ws['E1'] = 'Период'
        for user, vacations in data.items():
            width = widths.pop(0) - 1  # ширина    
            if width < 0:
                width = 0
            start = 'A' + start
            end = start[:1] + str(int(start[1:]) + width)
            full = start + ':' + end
            ws.merge_cells(full)
            ws[start] = user
            ws[start].fill = PatternFill("solid", "ABCDEF")
            ws[start].alignment = Alignment(horizontal="center", vertical="center")
            i = 0
            for vac in vacations:
                numbr = start.replace('A', 'B')[:1] + str(int(start.replace('A', 'B')[1:]) + i)
                ws[numbr].value = str(vac.day_start)[:10]
                numbr = start.replace('A', 'C')[:1] + str(int(start.replace('A', 'C')[1:]) + i)
                ws[numbr].value = str(vac.day_end)[:10]
                numbr = start.replace('A', 'D')[:1] + str(int(start.replace('A', 'D')[1:]) + i)
                ws[numbr].value = vac.how_long
                numbr = start.replace('A', 'E')[:1] + str(int(start.replace('A', 'E')[1:]) + i)
                ws[numbr].value = vac.year
                i += 1
            start = str(int(end[1:]) + 1)


        wb['Лист1'].column_dimensions['A'].width = 28
        wb['Лист1'].column_dimensions['B'].width = 15
        wb['Лист1'].column_dimensions['C'].width = 15
        wb['Лист1'].column_dimensions['D'].width = 15
        wb['Лист1'].column_dimensions['E'].width = 10
    
        wb.save(sample)
        pythoncom.CoInitializeEx(0)
        excel = client.Dispatch("Excel.Application")
        excel.Visible = False
        # Read Excel File
        sheets = excel.Workbooks.Open(sample)
        work_sheets = sheets.Worksheets[0]
        # Convert into PDF File
        pdf_file_name = (settings.BASE_DIR + f"/posts/static/users/vacations_{otd}_{year}.pdf")
        work_sheets.ExportAsFixedFormat(0, pdf_file_name)
        sheets.Close(True)

        return redirect('vac_2', year, otd)
    except:
        return redirect('vac_2', year, otd)
'''

def full_year(year):
    month_all = {
        1:'Январь', 2:'Февраль', 3:'Март',
        4:'Апрель', 5:'Май', 6:'Июнь',
        7:'Июль', 8:'Август', 9:'Сентябрь',
        10:'Октябрь', 11:'Ноябрь', 12:'Декабрь',
        }
    month_new = {}
    for i in range(1, 13):
        day = dt.datetime.today().replace(
            year=year,
            month=i,
            day=1
            )
        mnth_strt_d = day.replace(month=i, day=1).weekday()
        days_in_month = monthrange(year, i)[1]
        weeks, k = {}, 1
        weeks[k] = []
        #добавляем пустые клетки в начале месяца, если он начался не с понедельника
        [weeks[k].append('') for j in range(mnth_strt_d) if mnth_strt_d != 0]
        #заполняем месяц
        for j in range(1, days_in_month + 1):
            if mnth_strt_d < 7:
                weeks[k].append(j)
                mnth_strt_d += 1
            else:
                k += 1
                weeks[k] = []
                weeks[k].append(j)
                mnth_strt_d = 1
        #добавляем пустые клетки в конце месяца, если он закончился не в воскресенье    
        [weeks[k].append('') for j in range(mnth_strt_d, 7)]
        month_new[month_all[i]] = weeks
    return month_new

def vacations_start(request):
    try:
        otd_id = User_info.objects.filter(user_id=request.user.id)[0].otd_number_id
        otd = int(Unit.objects.filter(id=otd_id)[0].description)
    except:
        otd = 0
    year = dt.datetime.today().year
    bosses = {'Николай Емельянов': [305, 306, 307]}
    return render(
        request,
        'vacations_all.html',
        {**rights(request), 'otd': otd, 'bosses': bosses, 'year': year, },
    )